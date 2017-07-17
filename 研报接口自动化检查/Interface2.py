# -*- coding: UTF-8 -*-
from xml.dom import minidom
import xlrd
import openpyxl
import requests
import json
import sys
import HTMLParser
import os
import re
import codecs
import time
import datetime

reload(sys)
sys.setdefaultencoding('utf-8')

class OptionExcelData(object):
    """对Excel进行操作，包括读取请求参数，和填写操作结果"""
    def __init__(self, excelFile,excelPath=''):
        self.excelFile = excelFile
        self.excelPath = excelPath
        self.caseList = []

    """
    传入：传入用例Excel名称
    返回：[],其中元素为{},每个{}包含行号、城市、国家和期望结果的键值对
    """
    def getCaseList(self,excelFile,excelPath=''):
        readExcel = xlrd.open_workbook(fileName)                            #读取指定的Excel
        try:
            table = readExcel.sheet_by_index(0)                             #获取Excel的第一个sheet
            trows = table.nrows                                             #获取Excel的行数
            for n in range(1,trows):
                tmpdict = {}                                                #把一行记录写进一个{}
                tmpdict['id'] = n                                           #n是Excel中的第n行
                tmpdict['CityName'] = table.cell(n,2).value
                tmpdict['CountryName'] = table.cell(n,3).value
                tmpdict['Rspect'] = table.cell(n,4).value
                self.caseList.append(tmpdict)
        except Exception, e:
            raise
        finally:
            pass
        return self.caseList

    """
    传入：请求指定字段结果，是否通过，响应时间
    返回：
    """
    def writeCaseResult(self,resultBody,isSuccess,respTime,\
        excelFile,theRow,theCol=5):
        writeExcel = openpyxl.load_workbook(excelFile)                      #加载Excel，后续写操作
        try:
            wtable = writeExcel.get_sheet_by_name('Sheet1')                 #获取名为Sheet1的sheet
            wtable.cell(row=theRow+1,column=theCol+1).value = resultBody    #填写实际值
            wtable.cell(row=theRow+1,column=theCol+2).value = isSuccess     #填写是否通过
            wtable.cell(row=theRow+1,column=theCol+3).value = respTime      #填写响应时间
            writeExcel.save(excelFile)
        except Exception, e:
            raise
        finally:
            pass


class GetWeather(object):
    """获取天气的http请求"""
    def __init__(self, serviceUrl,requestBody,headers):
        self.serviceUrl = serviceUrl
        self.requestBody = requestBody
        self.headers = headers
        self.requestResult = {}

    """
    传入：请求地址，请求体，请求头
    返回：返回{}，包含响应时间和请求结果的键值对
    """
    def getWeath(self,serviceUrl,requestBody,headers):
        timebefore = time.time()                                            #获取请求开始的时间，不太严禁
        tmp = requests.post(serviceUrl,data=requestBody,\
            headers=headers)
        timeend = time.time()                                               #获取请求结束的时间
        tmptext = tmp.text
        self.requestResult['text'] = tmptext                                #记录响应回来的内容
        self.requestResult['time'] = round(timeend - timebefore,2)          #计算响应时间
        return self.requestResult

class XmlReader:
    """操作XML文件"""
    def __init__(self,testFile,testFilePath=''):
        self.fromXml = testFile
        self.xmlFilePath = testFilePath
        self.resultList = []

    def writeXmlData(self,resultBody,testFile,testFilePath=''):
        tmpXmlFile = codecs.open(testFile,'w','utf-16')                     #新建xml文件
        tmpLogFile = codecs.open(testFile+'.log','w','utf-16')              #新建log文件

        tmp1 = re.compile(r'\<.*?\>')                                     #生成正则表达式:<*?>
        tmp2 = tmp1.sub('',resultBody['text'])                              #替换相应结果中的<*?>
        html_parser = HTMLParser.HTMLParser()
        xmlText = html_parser.unescape(tmp2)                                #转换html编码

        try:
            tmpXmlFile.writelines(xmlText.strip())                          #去除空行并写入xml
            tmpLogFile.writelines('time: '+\
                str(resultBody['time'])+'\r\n')                             #把响应时间写入log
            tmpLogFile.writelines('text: '+resultBody['text'].strip())      #把请求回来的文本写入log
        except Exception, e:
            raise
        finally:
            tmpXmlFile.close()
            tmpLogFile.close()

    """返回一个list"""
    def readXmlData(self,testFile,testFilePath=''):
        tmpXmlFile = minidom.parse(testFile)
        root = tmpXmlFile.documentElement
        tmpValue = root.getElementsByTagName('City')[0].\
        childNodes[0].data
        return tmpValue                                                     #获取特定字段并返回结果，此处选取Status


if __name__ == '__main__':

    requesturl = 'http://www.webservicex.net/globalweather.asmx/GetCitiesByCountry'
    requestHeadrs = {"Content-Type":"application/x-www-form-urlencoded"}
    fileName = u'用例内容.xlsx'

    ed = OptionExcelData(fileName)
    testCaseList = ed.getCaseList(ed.excelFile)

    for caseDict in testCaseList:
        caseId = caseDict['id']
        cityName = caseDict['CityName']
        countryName = caseDict['CountryName']
        rspect = caseDict['Rspect']
        #requestBody = 'CityName='+cityName+'&CountryName='+countryName
        requestBody = 'CountryName='+countryName

        getWeather = GetWeather(requesturl,requestBody,requestHeadrs)
        #获取请求结果
        tmpString = getWeather.getWeath(getWeather.serviceUrl,\
            getWeather.requestBody,getWeather.headers)

        xd = XmlReader(str(caseId) + '.xml')
        print str(caseId) + '.xml'
        #把请求内容写入xml和log
        xd.writeXmlData(tmpString,xd.fromXml)
        response = xd.readXmlData(str(caseId) + '.xml')
        respTime = tmpString['time']
        if response == rspect:
            theResult = 'Pass'
        else:
            theResult = 'False'
        ed.writeCaseResult(response,theResult,respTime,fileName,caseId)